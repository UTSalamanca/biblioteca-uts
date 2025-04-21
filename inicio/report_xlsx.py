from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Side, Border
import os, re
from django.conf import settings
from django.http import HttpResponse

def insert_header_image(sheet):
    """
    Inserta una imagen de encabezado en la hoja del libro de Excel.
    """
    # Ruta de la imagen
    imagen_path = os.path.join(settings.BASE_DIR, 'inicio', 'static', 'img', 'header_new_uts.jpg')

    # Verifica si la imagen existe
    if not os.path.exists(imagen_path):
        raise FileNotFoundError(f"La imagen no existe en la ruta: {imagen_path}")

    # Inserta la imagen en el Excel
    img = Image(imagen_path)
    img.width = 700
    img.height = 100
    sheet.add_image(img, "B2")  # Inserta la imagen en la celda B2

def reporte_info(sheet, data):
    # Unión de celdas
    sheet.merge_cells('A11:H11')
    sheet['A11'].font = Font(color = 'FFFFFF', bold=True, size=12)
    sheet['A11'].fill = PatternFill('solid', start_color="d20606")
    sheet['A11'] = 'CONTROL DE REPORTES MENSUALES DE SERVICIOS: BIBLIOTECA.'

    sheet.merge_cells('A12:H12')
    sheet['A12'].font = Font(color = '000000', bold=True, size=12)
    sheet['A12'].fill = PatternFill('solid', start_color="d3c905")
    sheet['A12'] = 'CONSULTAS  EN EL CICLO DEL MES DE: ' + data['ciclo']

# def headers_by_tabla():
def get_borders(tipo):
    all_border = Border(
        left=Side(style="thin"),   # Borde izquierdo delgado
        top=Side(style="thin"),    # Borde superior delgado
        right=Side(style="thin"),   # Borde derecho delgado
        bottom=Side(style="thin")  # Borde inferior delgado
    )
    right_border = Border(
        top=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thin")
    )
    left_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        bottom=Side(style="thin")
    )
    top_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin")
    )
    bottom_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thin")
    )
    if tipo == 'all':
        regresa = all_border
    elif tipo == 'right':
        regresa = right_border
    elif tipo == 'left':
        regresa = left_border
    elif tipo == 'top':
        regresa = top_border
    elif tipo == 'bottom':
        regresa = bottom_border
    
    return regresa

# Crea la tabla de acervo
def table_acervo(sheet, data):
    """Función para la cración de registro para acervo

    Args:
        sheet (object): Instancia de Workbook
        data (array): Arreglo con la información repocilada
    """
    # configuraciones reusables reusables
    centrado = Alignment(horizontal='center', vertical='center')
    celdas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for c in celdas:
        sheet[f"{c}11"].fill = PatternFill('solid', start_color="0060df")
        sheet[f"{c}11"].font = Font(color = 'ffffff', bold=True, size=12)
        sheet[f"{c}11"].alignment = centrado
        sheet[f"{c}12"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}12"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{c}13"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}13"].font = Font(color = '000000', bold=True, size=12)
    # Unión de celdas
    sheet.merge_cells('A11:F11')
    sheet['A11'] = 'REPORTE GENERAL DE ACERVO BIBLIOGRÁFICO: ' + data['ciclo']

    # Crea encabezados de la tabla
    sheet['A12'].alignment = centrado
    sheet.merge_cells('A12:A13')
    sheet['A12'].border = get_borders('top')
    sheet['A13'].border = get_borders('bottom')
    sheet['A12'] = "No."
    
    # Agranda el tamaño de la celda A
    sheet.column_dimensions['A'].width = 10
    # Agranda el tamaño de la celda B
    sheet.column_dimensions['B'].width = 30
    sheet.merge_cells('B12:B13')
    sheet['B12'].border = get_borders('all')
    sheet['B13'].border = get_borders('all')
    sheet['B12'] = 'ÁREA DE CONOCIMIENTO'
    sheet['B12'].alignment = centrado
    # Se asignan las llaves de los grupo a las celdas
    cont_cell = 14
    cont_id = 1
    for key in data['conteo_ejemplares']:
        sheet[f"A{cont_cell}"] = cont_id
        sheet[f"A{cont_cell}"].alignment = centrado
        sheet[f"B{cont_cell}"] = key
        sheet[f"B{cont_cell}"].alignment = centrado
        cont_cell += 1
        cont_id += 1
    # Se asignan los valores en las celdas siguientes

    # ==> Se agregan titulos y volumenes de libros <==
    # Agranda el tamaño de la celda D
    sheet.column_dimensions['C'].width = 20
    # Unión de celdas
    sheet.merge_cells('C12:D12')
    # Se agregan bordes
    sheet['C12'].border = get_borders('left')
    sheet['D12'].border = get_borders('right')
    sheet['C12'] = 'LIBROS'
    sheet['C12'].alignment = centrado
    sheet['C13'].border = get_borders('all')
    sheet['C13'] = 'No.TITULOS'
    sheet['C13'].alignment = centrado

    cont_cell = 14
    totalizador_lib1 = 0
    for ord_libros in data['conteo_ejemplares']:
        if ord_libros in data['cantidad_libro']:
            sheet[f"C{cont_cell}"] = data['cantidad_libro'][ord_libros]
            totalizador_lib1 += data['cantidad_libro'][ord_libros]
        else:
            sheet[f"C{cont_cell}"] = 0
        sheet[f"C{cont_cell}"].alignment = centrado
        cont_cell += 1

    # Asigna número de volumenes
    sheet.column_dimensions['D'].width = 20
    sheet['D13'].border = get_borders('all')
    sheet['D13'] = 'No. VOLUMENES'
    sheet['D13'].alignment = centrado
    cont_cell = 14
    totalizador_lib2 = 0
    # for volms in data['volumenes_por_libro']:
    for ord_v_lib in data['conteo_ejemplares']:
        if ord_v_lib in data['volumenes_por_libro']:
            sheet[f"D{cont_cell}"] = data['volumenes_por_libro'][ord_v_lib]
            totalizador_lib2 += data['volumenes_por_libro'][ord_v_lib]
        else:
            sheet[f"D{cont_cell}"] = 0
        sheet[f"D{cont_cell}"].alignment = centrado
        cont_cell += 1

    # ==> Se agregan titulos y volumenes de "discos" <==
    sheet.column_dimensions['E'].width = 20
    # Unión de celdas
    sheet.merge_cells('E12:F12')
    # Se agregan bordes
    sheet['E12'].border = get_borders('left')
    sheet['F12'].border = get_borders('right')
    sheet['E12'] = 'CD-ROM'
    sheet['E12'].alignment = centrado
    sheet['E13'].border = get_borders('all')
    sheet['E13'] = 'No.TITULOS'
    sheet['E13'].alignment = centrado

    cont_cell = 14
    totalizador_disc1 = 0
    # for datas in data['cantidad_disco']:
    for ord_c_disc in data['conteo_ejemplares']:
        if ord_c_disc in data['cantidad_disco']:
            sheet[f"E{cont_cell}"] = data['cantidad_disco'][ord_c_disc]
            totalizador_disc1 += data['cantidad_disco'][ord_c_disc]
        else:
            sheet[f"E{cont_cell}"] = 0
        sheet[f"E{cont_cell}"].alignment = centrado
        cont_cell += 1
    # Asigna número de volumenes
    sheet.column_dimensions['F'].width = 20
    sheet['F13'].border = get_borders('all')
    sheet['F13'] = 'No. VOLUMENES'
    sheet['F13'].alignment = centrado

    cont_cell = 14
    totalizador_disc2 = 0
    # for volms in data['volumenes_por_disco']:
    for ord_v_disc in data['conteo_ejemplares']:
        if ord_v_disc in data['volumenes_por_disco']:
            sheet[f"F{cont_cell}"] = data['volumenes_por_disco'][ord_v_disc]
            totalizador_disc2 += data['volumenes_por_disco'][ord_v_disc]
        else:
            sheet[f"F{cont_cell}"] = 0
        sheet[f"F{cont_cell}"].alignment = centrado
        cont_cell += 1

    # ==> Se agregan titulos y volumenes de "revistas" <==
    sheet.column_dimensions['G'].width = 20
    # Unión de celdas
    sheet.merge_cells('G12:H12')
    # Se agregan bordes
    sheet['G12'].border = get_borders('left')
    sheet['H12'].border = get_borders('right')
    sheet['G12'] = 'REVISTAS'
    sheet['G12'].alignment = centrado
    sheet['G13'].border = get_borders('all')
    sheet['G13'] = 'No.TITULOS'
    sheet['G13'].alignment = centrado

    cont_cell = 14
    totalizador_revist1 = 0
    for ord_c_revist in data['conteo_ejemplares']:
        if ord_c_revist in data['cantidad_revista']:
            sheet[f"G{cont_cell}"] = data['cantidad_revista'][ord_c_revist]
            totalizador_revist1 += data['cantidad_revista'][ord_c_revist]
        else:
            sheet[f"G{cont_cell}"] = 0
        sheet[f"G{cont_cell}"].alignment = centrado
        cont_cell += 1
    # Asigna número de volumenes
    sheet.column_dimensions['H'].width = 20
    sheet['H13'].border = get_borders('all')
    sheet['H13'] = 'No. VOLUMENES'
    sheet['H13'].alignment = centrado

    cont_cell = 14
    totalizador_revist2 = 0
    for ord_v_revist in data['conteo_ejemplares']:
        if ord_v_revist in data['volumenes_por_revista']:
            sheet[f"H{cont_cell}"] = data['volumenes_por_revista'][ord_v_revist]
            totalizador_revist2 += data['volumenes_por_revista'][ord_v_revist]
        else:
            sheet[f"H{cont_cell}"] = 0
        sheet[f"H{cont_cell}"].alignment = centrado
        cont_cell += 1

    # ==> Tatalizadores
    for celda in celdas:
        # Color y estilo a la fila
        sheet[f"{celda}{cont_cell}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{cont_cell}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{cont_cell}"].border = get_borders('all')
        sheet[f"{celda}{cont_cell}"].alignment = centrado

    sheet[f"A{cont_cell}"] = 'Total'
    # Totalizador titulos libros
    sheet[f"C{cont_cell}"] = totalizador_lib1
    # Totalizador volumenes libros
    sheet[f"D{cont_cell}"] = totalizador_lib2
    # Totalizador titulos discos
    sheet[f"E{cont_cell}"] = totalizador_disc1
    # Totalizador volumenes discos
    sheet[f"F{cont_cell}"] = totalizador_disc2
    # Totalizador volumenes revistas
    sheet[f"G{cont_cell}"] = totalizador_revist1
    # Totalizador volumenes revistas
    sheet[f"H{cont_cell}"] = totalizador_revist2

    # ==> Tabla para adquisiciones acervo
    new_cell = cont_cell + 2
    # Unión de celdas
    sheet.merge_cells(f"A{new_cell}:H{new_cell}")
    # Da formato a las cabeceras
    for celda in celdas:
        sheet[f"{celda}{new_cell}"].fill = PatternFill('solid', start_color="0060df")
        sheet[f"{celda}{new_cell}"].font = Font(color = 'ffffff', bold=True, size=12)
        sheet[f"{celda}{new_cell}"].border = get_borders('all')
        sheet[f"{celda}{new_cell}"].alignment = centrado
        sheet[f"{celda}{new_cell + 1}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{new_cell + 1}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{new_cell + 1}"].border = get_borders('all')
        sheet[f"{celda}{new_cell + 1}"].alignment = centrado
        sheet[f"{celda}{new_cell + 2}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{new_cell + 2}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{new_cell + 2}"].border = get_borders('all')
        sheet[f"{celda}{new_cell + 2}"].alignment = centrado
    # Agrega titulos
    sheet[f"A{new_cell}"] = "ADQUISICIÓN DE ACERVO BIBLIOGRAFICO MES DE: " + data['ciclo']
     # Crea encabezados de la tabla
     # Titulo 1
    sheet[f"A{new_cell + 1}"].alignment = centrado
    sheet.merge_cells(f"A{new_cell + 1}:A{new_cell + 2}")
    sheet[f"A{new_cell + 1}"].border = get_borders('top')
    sheet[f"A{new_cell + 2}"].border = get_borders('bottom')
    sheet[f"A{new_cell + 1}"] = "No."
    # Titulo 2
    sheet.merge_cells(f"B{new_cell + 1}:B{new_cell + 2}")
    sheet[f"B{new_cell + 1}"].border = get_borders('top')
    sheet[f"B{new_cell + 1}"] = 'ÁREA DE CONOCIMIENTO'
    sheet[f"B{new_cell}"].alignment = centrado
    sheet[f"B{new_cell + 2}"].border = get_borders('bottom')
    # Titulo 3
    sheet.merge_cells(f"C{new_cell + 1}:D{new_cell + 1}")
    sheet[f"C{new_cell + 1}"] = "LIBROS"
    # Titulo 3.1
    sheet[f"C{new_cell + 2}"] = "NO.TÍTULOS"
    sheet[f"D{new_cell + 2}"] = "NO.VOLUMENES"
    # Titulo 4
    sheet.merge_cells(f"E{new_cell + 1}:F{new_cell + 1}")
    sheet[f"E{new_cell + 1}"] = "CD-ROM"    
    # Titulo 4.1
    sheet[f"E{new_cell + 2}"] = "NO.TÍTULOS"
    sheet[f"F{new_cell + 2}"] = "NO.VOLUMENES"
    # Titulo 5
    sheet.merge_cells(f"G{new_cell + 1}:H{new_cell + 1}")
    sheet[f"G{new_cell + 1}"] = "REVISTAS"    
    # Titulo 5.1
    sheet[f"G{new_cell + 2}"] = "NO.TÍTULOS"
    sheet[f"H{new_cell + 2}"] = "NO.VOLUMENES"
    # Se procesa la información
    contador = 1
    data_cell = new_cell + 3
    totalizador_lib1 = 0
    totalizador_lib2 = 0
    totalizador_disc1 = 0
    totalizador_disc2 = 0
    totalizador_revist1 = 0
    totalizador_revist2 = 0
    # Coloca nombre
    for area_c in data['conteo_ejemplares']:
        sheet[f"A{data_cell}"] = contador
        sheet[f"A{data_cell}"].alignment = centrado
        sheet[f"B{data_cell}"] = area_c
        sheet[f"B{data_cell}"].alignment = centrado
        vol_libros = data['adquisiciones']['volumen_libros'] if len(data['adquisiciones']['volumen_libros']) != 0 else 0
        cant_libros = data['adquisiciones']['cantidad_libros'] if len(data['adquisiciones']['cantidad_libros']) != 0 else 0
        vol_discos = data['adquisiciones']['volumen_discos'] if len(data['adquisiciones']['volumen_discos']) != 0 else 0
        cant_discos = data['adquisiciones']['cantidad_discos'] if len(data['adquisiciones']['cantidad_discos']) != 0 else 0
        vol_revistas = data['adquisiciones']['volumen_revistas'] if len(data['adquisiciones']['volumen_revistas']) != 0 else 0
        cant_revistas = data['adquisiciones']['cantidad_revistas'] if len(data['adquisiciones']['cantidad_revistas']) != 0 else 0
        # Libros
        if vol_libros != 0:
            dato_t = cant_libros.get(area_c) if cant_libros.get(area_c) != None else 0
            sheet[f"C{data_cell}"] = dato_t
            dato_v = vol_libros.get(area_c) if vol_libros.get(area_c) != None else 0
            sheet[f"D{data_cell}"] = dato_v
            # Totalizadores
            totalizador_lib1 += dato_v
            totalizador_lib2 += dato_t
        else:
            sheet[f"C{data_cell}"] = 0
            sheet[f"D{data_cell}"] = 0

        sheet[f"C{data_cell}"].alignment = centrado
        sheet[f"D{data_cell}"].alignment = centrado
        # Discos
        if vol_discos != 0:
            dato_t = cant_discos.get(area_c) if cant_discos.get(area_c) != None else 0
            sheet[f"E{data_cell}"] = dato_t
            dato_v = vol_discos.get(area_c) if vol_discos.get(area_c) != None else 0
            sheet[f"F{data_cell}"] = dato_v
            # Totalizadores
            totalizador_disc1 += dato_v
            totalizador_disc2 += dato_t
        else:
            sheet[f"E{data_cell}"] = 0
            sheet[f"F{data_cell}"] = 0
        
        sheet[f"E{data_cell}"].alignment = centrado
        sheet[f"F{data_cell}"].alignment = centrado
        # Revistas
        if vol_revistas != 0:
            dato_t = cant_revistas.get(area_c) if cant_revistas.get(area_c) != None else 0
            sheet[f"G{data_cell}"] = dato_t
            dato_v = vol_revistas.get(area_c) if vol_revistas.get(area_c) != None else 0
            sheet[f"H{data_cell}"] = dato_v
            # Totalizadores
            totalizador_revist1 += dato_v
            totalizador_revist2 += dato_t
        else:
            sheet[f"G{data_cell}"] = 0
            sheet[f"H{data_cell}"] = 0
        
        sheet[f"G{data_cell}"].alignment = centrado
        sheet[f"H{data_cell}"].alignment = centrado
        contador += 1
        data_cell += 1

    # Sección de totalizadores

    for celda in celdas:
        # Color y estilo a la fila
        sheet[f"{celda}{data_cell}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{data_cell}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{data_cell}"].border = get_borders('all')
        sheet[f"{celda}{data_cell}"].alignment = centrado

    sheet[f"A{data_cell}"] = 'Total'
    # Totalizador ejemplares libro
    sheet[f"C{data_cell}"] = totalizador_lib2
    # Totalizador volumenes libro
    sheet[f"D{data_cell}"] = totalizador_lib1
    # Totalizador ejemplares disco
    sheet[f"E{data_cell}"] = totalizador_disc2
    # Totalizador volumenes disco
    sheet[f"F{data_cell}"] = totalizador_disc1
    # Totalizador volumenes revista
    sheet[f"G{data_cell}"] = totalizador_revist2
    # Totalizador volumenes revista
    sheet[f"H{data_cell}"] = totalizador_revist1
        
def table_reporte_estadias(sheet, data):
    """Función para la creación de la tabla de reportes de estadías

    Args:
        sheet (object): Instancia de Workbook
        data (_type_): Arreglo con la información recopilada
    """
    # configuraciones reusables reusables
    centrado = Alignment(horizontal='center', vertical='center')
    # Agrega caracteristicas grupales
    celdas = ['A', 'B', 'C', 'D', 'E']
    for c in celdas:
        sheet[f"{c}11"].fill = PatternFill('solid', start_color="0060df")
        sheet[f"{c}11"].font = Font(color = 'ffffff', bold=True, size=12)
        sheet[f"{c}11"].border = get_borders('all')
        sheet[f"{c}11"].alignment = centrado
        sheet[f"{c}12"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}12"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{c}12"].border = get_borders('all')
        sheet[f"{c}12"].alignment = centrado
    # Unión de celdas
    sheet.merge_cells('A11:E11')
    sheet['A11'] = 'REPORTE GENERAL DOCUMENTOS DE ESTADÍA: ' + data['ciclo']
    # Ajuste de ancho de celda
    sheet.column_dimensions['C'].width = 70
    # Crea encabezados de la tabla
    sheet.row_dimensions[12].height = 25
    sheet['A12'] = "No."
    # 
    sheet['B12'] = 'Abreviatura'
    sheet['C12'] = 'Área de conocimiento'
    # Se agregan numero de reportes por carrera
    sheet.column_dimensions['D'].width = 30
    sheet['D12'] = "No. Reportes"
    # Se agregan numero de vistas de reportes por carrera
    sheet.column_dimensions['E'].width = 30
    sheet['E12'] = "Total de vistas"
    # Se agrega numéración de campos
    # Se agregan careras activas
    cont_cell = 13
    cont_id = 1
    totalizador1 = 0
    totalizador2 = 0
    for carrera in data['conc_carreras']:
        sheet[f"A{cont_cell}"] = cont_id
        sheet[f"A{cont_cell}"].alignment = centrado
        # Se agregan campos de abreviaciones y nombres de las carreras
        sheet[f"B{cont_cell}"] = carrera
        sheet[f"B{cont_cell}"].alignment = centrado
        sheet[f"C{cont_cell}"] = data['conc_carreras'][carrera]
        sheet[f"C{cont_cell}"].alignment = centrado
        # Se agregan numero de reportes por carrera
        coincidencia = True
        for repo in data['estadias_reportes']:
            if re.search(re.escape(carrera), repo, re.IGNORECASE):
                sheet[f"D{cont_cell}"] = data['estadias_reportes'][repo]
                # Suma para totalizador
                totalizador1 += data['estadias_reportes'][repo]
                coincidencia = True
                break
            coincidencia = False
        # Si no existio ninguna coincidencia marca en 0
        if not coincidencia:
            sheet[f"D{cont_cell}"] = 0
        # Centrado de contenido
        sheet[f"D{cont_cell}"].alignment = centrado
        # Se agregan el numero vistas de reportes por carrera
        por_carrea = {}
        concentrado = {}
        coincidencia = True
        for vista in data['cont_vistas_reporte']:
            grupo_carr = data['cont_vistas_reporte'][vista]
            if grupo_carr[0] in concentrado:
                ant = concentrado[grupo_carr[0]]
                concentrado[grupo_carr[0]] = ant + grupo_carr[1]
            else:
                concentrado[grupo_carr[0]] = grupo_carr[1]
        
        for conc in concentrado:
            if re.search(re.escape(carrera), conc, re.IGNORECASE):
                sheet[f"E{cont_cell}"] = concentrado[conc]
                totalizador2 += concentrado[conc]
                coincidencia = True
                break
            coincidencia = False
        if not coincidencia:
            sheet[f"E{cont_cell}"] = 0
        # Centrado de contenido
        sheet[f"E{cont_cell}"].alignment = centrado
        # Incremento de contadores                
        cont_cell += 1
        cont_id += 1

    # ==> Totalizadores
    for celda in celdas:
        sheet[f"{celda}{cont_cell}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{cont_cell}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{cont_cell}"].border = get_borders('all')
        sheet[f"{celda}{cont_cell}"].alignment = centrado

    sheet[f"A{cont_cell}"] = "Total"
    sheet[f"D{cont_cell}"] = totalizador1
    sheet[f"E{cont_cell}"] = totalizador2

    # ==> Tabla vistas por reporte
    celdas = ['A', 'B', 'C', 'D']
    new_cell = cont_cell + 3
    for c in celdas:
        sheet[f"{c}{new_cell - 1}"].fill = PatternFill('solid', start_color="0060df")
        sheet[f"{c}{new_cell - 1}"].font = Font(color = 'ffffff', bold=True, size=12)
        sheet[f"{c}{new_cell - 1}"].border = get_borders('all')
        sheet[f"{c}{new_cell}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}{new_cell}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{c}{new_cell}"].border = get_borders('all')
        sheet[f"{c}{new_cell}"].alignment = centrado

    sheet.merge_cells(f"A{new_cell - 1}:D{new_cell - 1}")
    sheet[f"A{new_cell - 1}"] = 'REPORTE GENERAL DE VISTAS POR PROYECTO: ' + data['ciclo']
    sheet[f"A{new_cell - 1}"].alignment = centrado
    # Aumenta el ancho de la fila
    sheet.row_dimensions[new_cell].height = 25
    # Header para numeración
    sheet[f"A{new_cell}"] = "No."
    # Nombre de proyecto
    sheet[f"B{new_cell}"] = "Nombre de proyecto"
    sheet[f"B{new_cell}"].alignment = centrado
    sheet.column_dimensions['B'].width = 30
    # Nombre de la carrea
    sheet[f"C{new_cell}"] = "Carrera relacionada"
    sheet[f"C{new_cell}"].alignment = centrado
    # Numero de vistas por proyecto
    sheet[f"D{new_cell}"] = "No. vistas"
    sheet[f"D{new_cell}"].alignment = centrado
    # Se agrega la información
    contador_vistas = 1
    for vista in data['cont_vistas_reporte']:
        carrera = data['cont_vistas_reporte'][vista][0]
        no_vistas = data['cont_vistas_reporte'][vista][1]
        # Nuevo numero de celda
        celda = new_cell + contador_vistas
        sheet[f"A{celda}"] = contador_vistas
        sheet[f"A{celda}"].alignment = centrado
        sheet[f"B{celda}"] = vista
        sheet[f"B{celda}"].alignment = centrado
        sheet[f"C{celda}"] = carrera
        sheet[f"C{celda}"].alignment = centrado
        sheet[f"D{celda}"] = no_vistas
        sheet[f"D{celda}"].alignment = centrado
        contador_vistas += 1
    # Se agrega filtro en celda finales
    sheet.auto_filter.ref = f"B{new_cell}:C{celda}"
    # Suma de datos en columna
    sheet[f"D{celda + 1}"] = f"=SUM(D{cont_cell + 3}:D{celda})"
    # Numero de vistas por proyecto
    sheet[f"A{celda + 1}"] = "Total"
    sheet[f"A{celda + 1}"].alignment = centrado
    for c in celdas:
        sheet[f"{c}{celda + 1}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}{celda + 1}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{c}{celda + 1}"].border = get_borders('all')
        sheet[f"{c}{celda + 1}"].alignment = centrado

def table_prestamos(sheet, data):
    """Crea tabla de prestamos

    Args:
        sheet (object): Instancia de Workbook
        data (array): Arreglo con la información formateada.
    """
        # configuraciones reusables reusables
    centrado = Alignment(horizontal='center', vertical='center')
    # Agrega caracteristicas grupales
    celdas = ['A', 'B', 'C', 'D', 'E', 'F']
    for c in celdas:
        sheet[f"{c}11"].fill = PatternFill('solid', start_color="0060df")
        sheet[f"{c}11"].font = Font(color = 'ffffff', bold=True, size=12)
        sheet[f"{c}11"].border = get_borders('all')
        sheet[f"{c}11"].alignment = centrado
        sheet[f"{c}12"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}12"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{c}12"].border = get_borders('all')
        sheet[f"{c}12"].alignment = centrado
        sheet[f"{c}13"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{c}13"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{c}13"].border = get_borders('all')
        sheet[f"{c}13"].alignment = centrado
    # Unión de celdas
    sheet.merge_cells('A11:F11')
    sheet['A11'] = 'REPORTE DE PRESTAMOS: ' + data['ciclo']
    # Ajuste de ancho de celda
   # Crea encabezados de la tabla
    sheet['A12'].alignment = centrado
    sheet.merge_cells('A12:A13')
    sheet['A12'].border = get_borders('top')
    sheet['A13'].border = get_borders('bottom')
    sheet['A12'] = "No."
    
    # Agranda el tamaño de la celda A
    sheet.column_dimensions['A'].width = 10
    # Agranda el tamaño de la celda B
    sheet.column_dimensions['B'].width = 30
    sheet.merge_cells('B12:B13')
    sheet['B12'].border = get_borders('all')
    sheet['B13'].border = get_borders('all')
    sheet['B12'] = 'ÁREA DE CONOCIMIENTO'
    sheet['B12'].alignment = centrado

    # ==> Se agregan titulos y volumenes de libros <==
    # Agranda el tamaño de la celda D
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    # Unión de celdas
    sheet.merge_cells('C12:D12')
    sheet.merge_cells('E12:F12')
    # Sección prestamo interno
    sheet['C12'].border = get_borders('left')
    sheet['D12'].border = get_borders('right')
    sheet['C12'] = 'INTERNO'
    sheet['C12'].alignment = centrado
    sheet['C13'].border = get_borders('all')
    sheet['C13'] = 'PRESTADOS'
    sheet['C13'].alignment = centrado
    sheet['D13'].border = get_borders('all')
    sheet['D13'] = 'NO DEVUELTOS'
    sheet['D13'].alignment = centrado
    # Sección prestamo externo
    sheet['E12'].border = get_borders('left')
    sheet['F12'].border = get_borders('right')
    sheet['E12'] = 'EXTERNO'
    sheet['E12'].alignment = centrado
    sheet['E13'].border = get_borders('all')
    sheet['E13'] = 'PRESTADOS'
    sheet['E13'].alignment = centrado
    sheet['F13'].border = get_borders('all')
    sheet['F13'] = 'NO DEVUELTOS'
    sheet['F13'].alignment = centrado

    # Se agrega información
    # Se asignan las llaves de los grupo a las celdas
    cont_cell = 14
    tot_prestamo_int = 0
    tot_noDevuelto_int = 0
    tot_prestamo_ext = 0
    tot_noDevuelto_ext = 0
    cont_id = 1
    for key in data['conteo_ejemplares']:
        sheet[f"A{cont_cell}"] = cont_id
        sheet[f"A{cont_cell}"].alignment = centrado
        sheet[f"B{cont_cell}"] = key
        sheet[f"B{cont_cell}"].alignment = centrado
        # Agregan prestamos internos
        if len(data['prestamo_conc']['prestados_int']) != 0:
            if key in data['prestamo_conc']['prestados_int']:
                sheet[f"C{cont_cell}"] = data['prestamo_conc']['prestados_int'][key]
                tot_prestamo_int += data['prestamo_conc']['prestados_int'][key]
            else:
                sheet[f"C{cont_cell}"] = 0
        else:
            sheet[f"C{cont_cell}"] = 0
        sheet[f"C{cont_cell}"].alignment = centrado

        if len(data['prestamo_conc']['noDevueltos_int']) != 0:
            if key in data['prestamo_conc']['noDevueltos_int']:
                sheet[f"D{cont_cell}"] = data['prestamo_conc']['noDevueltos_int'][key]
                tot_noDevuelto_int += data['prestamo_conc']['noDevueltos_int'][key]
            else:
                sheet[f"D{cont_cell}"] = 0
        else:
            sheet[f"D{cont_cell}"] = 0
        sheet[f"D{cont_cell}"].alignment = centrado
        if len(data['prestamo_conc']['prestados_ext']) > 0:
            if key in data['prestamo_conc']['prestados_ext']:
                sheet[f"E{cont_cell}"] = data['prestamo_conc']['prestados_ext'][key]
                tot_prestamo_ext += data['prestamo_conc']['prestados_ext'][key]
            else:
                sheet[f"E{cont_cell}"] = 0
        else:
            sheet[f"E{cont_cell}"] = 0
        sheet[f"E{cont_cell}"].alignment = centrado

        if len(data['prestamo_conc']['noDevueltos_ext']) != 0:
            if key in data['prestamo_conc']['noDevueltos_ext']:
                sheet[f"F{cont_cell}"] = data['prestamo_conc']['noDevueltos_ext'][key]
                tot_noDevuelto_ext += data['prestamo_conc']['noDevueltos_ext'][key]
            else:
                sheet[f"F{cont_cell}"] = 0
        else:
            sheet[f"F{cont_cell}"] = 0
        sheet[f"F{cont_cell}"].alignment = centrado

        cont_cell += 1
        cont_id += 1
    for celda in celdas:
        sheet[f"{celda}{cont_cell}"].fill = PatternFill('solid', start_color="a0aab9")
        sheet[f"{celda}{cont_cell}"].font = Font(color = '000000', bold=True, size=12)
        sheet[f"{celda}{cont_cell}"].border = get_borders('all')
        sheet[f"{celda}{cont_cell}"].alignment = centrado

    sheet[f"A{cont_cell}"] = 'Total'
    sheet[f"C{cont_cell}"] = tot_prestamo_int
    sheet[f"D{cont_cell}"] = tot_noDevuelto_int
    sheet[f"E{cont_cell}"] = tot_prestamo_ext
    sheet[f"F{cont_cell}"] = tot_noDevuelto_ext

def create_excel(data):
    """Crea un archivo Excel con una imagen de encabezado.

    Args:
        data (array): Arreglo con la información necesaria para la creación del archivo Excel

    Returns:
        object: Instancia de Workbook
    """
    # Crear el libro y la hoja
    book = Workbook()
    book.remove(book['Sheet']) # Se remueve la hoja que se activa por degfault

    acervo = book.create_sheet(title="Acervo")
    # Inserta la imagen en el encabezado
    insert_header_image(acervo)
    # Inserta información
    table_acervo(acervo, data)
    # Creación tabla de estadías 
    estadias = book.create_sheet(title="Estadías")
    insert_header_image(estadias)
    table_reporte_estadias(estadias, data)
    # Tablas de prestamos
    prestamos = book.create_sheet(title="Prestamos")
    insert_header_image(prestamos)
    table_prestamos(prestamos, data)

    return book

def generate_report(data):
    try:
        # Crear el archivo Excel
        book = create_excel(data)
        # Configurar la respuesta HTTP para la descarga
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="Reporte mensual '+ data['ciclo'] +'.xlsx"'

        # Guardar el archivo Excel directamente en la respuesta
        book.save(response)

        return response
    except FileNotFoundError as e:
        # Manejo específico para errores relacionados con la imagen
        return HttpResponse(str(e), status=404)
    except Exception as e:
        # Manejo genérico de errores
        return HttpResponse(f"Error al generar el reporte: {str(e)}", status=500)
